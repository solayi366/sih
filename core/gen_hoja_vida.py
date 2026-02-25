"""
SIH — Generador de Hoja de Vida en Excel
Replica exactamente el formato de HOJA_DE_VIDA_06FAC02.xlsx
Uso: python3 gen_hoja_vida.py --demo       (genera un archivo de prueba)
     Se importa como módulo desde exportarController.php vía CLI
"""

import sys
import json
import openpyxl
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill)
from openpyxl.utils import get_column_letter
from datetime import datetime


# ════════════════════════════════════════════════════════════════════════════
#  ESTILOS REUTILIZABLES
# ════════════════════════════════════════════════════════════════════════════
def _font(bold=False, size=11, color=None):
    kw = dict(name='Calibri', bold=bold, size=size)
    if color:
        kw['color'] = color
    return Font(**kw)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(rgb):
    return PatternFill(fill_type='solid', fgColor=rgb)


# ════════════════════════════════════════════════════════════════════════════
#  FUNCIONES AUXILIARES DE ESCRITURA
# ════════════════════════════════════════════════════════════════════════════
def w(ws, coord, value, bold=False, size=11, halign='left'):
    """Escribe un valor con fuente Calibri."""
    c = ws[coord]
    c.value = value
    c.font = _font(bold=bold, size=size)
    c.alignment = _align(h=halign)

def wlabel(ws, coord, text):
    """Etiqueta de campo: bold size 14."""
    w(ws, coord, text, bold=True, size=14)

def wval(ws, coord, text):
    """Valor de campo: normal size 11."""
    w(ws, coord, text, bold=False, size=11)

def wsection(ws, coord, text):
    """Título de sección (CPU, MONITOR…): bold size 11."""
    w(ws, coord, text, bold=True, size=11)

def wsubfield(ws, coord, text):
    """Sub-etiqueta dentro de sección (MARCA, REFERENCIA…): size 10."""
    c = ws[coord]
    c.value = text
    c.font = _font(bold=False, size=10)
    c.alignment = _align()

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ════════════════════════════════════════════════════════════════════════════
#  HOJA "Inventario" — layout fijo replicando el template
# ════════════════════════════════════════════════════════════════════════════
def build_inventario(ws, data: dict):
    """
    data: {
      nombre_equipo, usuario, dependencia, fecha, tipo_equipo,
      responsable,
      cpu: {marca, referencia, serial, windows, service_pack, office, ip, mac,
             procesador, ram, disco_duro, cd},
      lector:  {marca, referencia, serie, serie_envia},
      monitor: {marca, referencia, serie},
      mouse:   {marca, referencia, serie},
      teclado: {marca, referencia, serie},
      ups:     {marca, referencia, serie},
      tarjeta_red: {marca, referencia, serie, mac},
      impresora:   {marca, referencia, serie, ip},
      telefono:    {referencia, serie, mac},
      observaciones,
      base_lector,
    }
    Para tablets/portátiles data tiene claves diferentes — se ajusta abajo.
    """

    # ── Anchos y altos ───────────────────────────────────────────────────
    col_w = {'A': 5.44, 'B': 14.89, 'C': 35.22, 'D': 4.0,
             'E': 12.33, 'F': 11.66, 'G': 22.55}
    for col, wd in col_w.items():
        ws.column_dimensions[col].width = wd

    row_h = {2:21, 3:18, 4:9.75, 5:20.25, 6:20.25, 7:20.25, 8:9.75,
             9:16.5, 10:14.25}
    for r in range(11, 45):
        row_h[r] = 20.25
    for r in [37,38,39,40,41]: row_h[r] = 21
    for r in [42,43]: row_h[r] = 18.75
    row_h[44] = 15
    for r, h in row_h.items():
        ws.row_dimensions[r].height = h

    # ── Fila 2: título ───────────────────────────────────────────────────
    ws['B2'] = 'INVENTARIO DE EQUIPOS'
    ws['B2'].font = _font(bold=True, size=16)
    merge(ws, 2, 2, 2, 7)  # B2:G2

    # ── Fila 3-4: fecha ──────────────────────────────────────────────────
    wlabel(ws, 'E3', 'Fecha:')
    merge(ws, 3, 5, 4, 5)  # E3:E4
    ws['F4'] = data.get('fecha', datetime.today())
    ws['F4'].font = _font(bold=True, size=14)
    merge(ws, 4, 6, 4, 7)  # F4:G4

    # ── Filas 5-7: datos generales ───────────────────────────────────────
    wlabel(ws, 'B5', 'Usuario:')
    wval(ws, 'C5', data.get('usuario', ''))

    wlabel(ws, 'B6', 'Nombre equipo: ')
    wval(ws, 'C6', data.get('nombre_equipo', ''))
    wlabel(ws, 'E6', 'Tipo de Equipo')
    merge(ws, 6, 5, 6, 7)

    wlabel(ws, 'B7', 'Dependencia:')
    ws['C7'].value = data.get('dependencia', '')
    ws['C7'].font = _font(bold=True, size=11)
    ws['E7'].value = data.get('tipo_equipo_label', '')
    ws['E7'].font = _font(bold=False, size=14)
    merge(ws, 7, 5, 7, 7)

    # ════ COLUMNA IZQUIERDA: CPU / info principal ════════════════════════
    is_tablet = data.get('es_tablet', False)
    cpu = data.get('cpu', {})

    if is_tablet:
        seccion_label = 'TABLET'
    elif 'laptop' in data.get('tipo_raw', '').lower() or 'portatil' in data.get('tipo_raw','').lower():
        seccion_label = 'PORTATIL'
    else:
        seccion_label = 'CPU'

    wsection(ws, 'B10', seccion_label)
    merge(ws, 10, 2, 10, 3)

    row = 11
    campos_cpu_pc = [
        ('MARCA',      cpu.get('marca', '')),
        ('REFERENCIA', cpu.get('referencia', '')),
        ('SERIAL',     cpu.get('serial', '')),
    ]
    if is_tablet:
        campos_cpu_pc += [
            ('MODELO',         cpu.get('modelo', '')),
            ('SERIAL ENVIA',   cpu.get('serial_envia', '')),
            ('SISTEMA OPERATIVO', cpu.get('so', '')),
            ('VERSION',        cpu.get('version', '')),
            ('IP',             cpu.get('ip', '')),
            ('MAC',            cpu.get('mac', '')),
            ('PROCESADOR',     cpu.get('procesador', '')),
            ('ALMACENAMIENTO', cpu.get('almacenamiento', '')),
        ]
    else:
        campos_cpu_pc += [
            ('WINDOWS',      cpu.get('windows', '')),
            ('SERVICE PACK',  cpu.get('service_pack', '')),
            ('OFFICE',        cpu.get('office', '')),
            ('IP',            cpu.get('ip', '')),
            ('MAC',           cpu.get('mac', '')),
            ('PROCESADOR',    cpu.get('procesador', '')),
            ('RAM',           cpu.get('ram', '')),
            ('DISCO DURO',    cpu.get('disco_duro', '')),
            ('CD',            cpu.get('cd', '')),
        ]

    for label, val in campos_cpu_pc:
        wsubfield(ws, f'B{row}', label)
        wval(ws, f'C{row}', str(val) if val else '')
        row += 1

    # TECLADO (solo PC/laptop)
    if not is_tablet:
        teclado = data.get('teclado', {})
        wsection(ws, f'B{row}', 'TECLADO')
        merge(ws, row, 2, row, 3)
        row += 1
        for lbl, val in [('MARCA', teclado.get('marca','')),
                          ('REFERENCIA', teclado.get('referencia','')),
                          ('SERIE', teclado.get('serie',''))]:
            wsubfield(ws, f'B{row}', lbl)
            wval(ws, f'C{row}', str(val) if val else '')
            row += 1

    # TARJETA RED INALAMBRICA
    tred = data.get('tarjeta_red', {})
    if tred:
        wsection(ws, f'B{row}', 'TARJETA RED INALAMBRICA')
        merge(ws, row, 2, row, 3)
        row += 1
        tr_campos = [('MARCA', tred.get('marca','')),
                     ('REFERENCIA', tred.get('referencia','')),
                     ('SERIE', tred.get('serie','')),
                     ('MAC ', tred.get('mac',''))]
        for lbl, val in tr_campos:
            wsubfield(ws, f'B{row}', lbl)
            wval(ws, f'C{row}', str(val) if val else '')
            row += 1

    # TELEFONO
    tel = data.get('telefono', {})
    if tel:
        wsection(ws, f'B{row}', 'TELEFONO ')
        merge(ws, row, 2, row, 3)
        row += 1
        for lbl, val in [('REFERENCIA', tel.get('referencia','')),
                         ('SERIE T', tel.get('serie','')),
                         ('MAC T', tel.get('mac',''))]:
            wsubfield(ws, f'B{row}', lbl)
            wval(ws, f'C{row}', str(val) if val else '')
            row += 1

    col_izq_fin = row  # hasta donde llegó la columna izquierda

    # ════ COLUMNA DERECHA: periféricos ══════════════════════════════════
    # LECTOR — arranca en fila 10
    lector = data.get('lector', {})
    wsection(ws, 'E10', 'LECTOR')
    merge(ws, 10, 5, 10, 7)

    ws['E11'] = 'MARCA';      ws['F11'] = 'REFERENCIA'; ws['G11'] = 'SERIE'
    for col in ['E11','F11','G11']:
        ws[col].font = _font(size=11)
    ws['E12'] = str(lector.get('marca',''));      ws['E12'].font = _font(size=11)
    ws['F12'] = str(lector.get('referencia','')); ws['F12'].font = _font(size=11)
    ws['G12'] = str(lector.get('serie',''));       ws['G12'].font = _font(size=11)
    merge(ws, 12, 6, 12, 7)

    # SERIE ENVIA del lector
    ws['E13'] = 'SERIE ENVIA'
    ws['E13'].font = _font(size=11)
    ws['F13'] = str(lector.get('serie_envia', ''))
    ws['F13'].font = _font(size=11)
    merge(ws, 13, 6, 13, 7)

    # BASE DE LECTOR
    ws['E14'] = 'BASE LECTOR'
    ws['E14'].font = _font(bold=True, size=11)

    base_lect = data.get('base_lector', {})
    ws['E15'] = 'MARCA';      ws['F15'] = str(base_lect.get('marca',''))
    ws['E16'] = 'REFERENCIA'; ws['F16'] = str(base_lect.get('referencia',''))
    ws['E17'] = 'SERIE';      ws['F17'] = str(base_lect.get('serie',''))
    ws['E18'] = 'SERIE ENVIA';ws['F18'] = str(base_lect.get('serie_envia',''))
    for r in range(15, 19):
        for col in ['E','F']:
            ws[f'{col}{r}'].font = _font(size=11)
        merge(ws, r, 6, r, 7)

    # MONITOR
    monitor = data.get('monitor', {})
    wsection(ws, 'E20', 'MONITOR')
    merge(ws, 20, 5, 20, 7)
    ws['E21'] = 'MARCA';      ws['F21'] = str(monitor.get('marca',''))
    ws['E22'] = 'REFERENCIA'; ws['F22'] = str(monitor.get('referencia',''))
    ws['E23'] = 'SERIE';      ws['F23'] = str(monitor.get('serie',''))
    for r in range(21, 24):
        for col in ['E','F']:
            ws[f'{col}{r}'].font = _font(size=11)
        merge(ws, r, 6, r, 7)

    # MOUSE
    mouse = data.get('mouse', {})
    wsection(ws, 'E25', 'MOUSE')
    merge(ws, 25, 5, 25, 7)
    ws['E26'] = 'MARCA';      ws['F26'] = str(mouse.get('marca',''))
    ws['E27'] = 'REFERENCIA'; ws['F27'] = str(mouse.get('referencia',''))
    ws['E28'] = 'SERIE';      ws['F28'] = str(mouse.get('serie',''))
    for r in range(26, 29):
        for col in ['E','F']:
            ws[f'{col}{r}'].font = _font(size=11)
        merge(ws, r, 6, r, 7)

    # UPS
    ups = data.get('ups', {})
    wsection(ws, 'E30', 'UPS')
    merge(ws, 30, 5, 30, 7)
    ws['E31'] = 'MARCA';      ws['F31'] = str(ups.get('marca',''))
    ws['E32'] = 'REFERENCIA'; ws['F32'] = str(ups.get('referencia',''))
    ws['E33'] = 'SERIE';      ws['F33'] = str(ups.get('serie',''))
    for r in range(31, 34):
        for col in ['E','F']:
            ws[f'{col}{r}'].font = _font(size=11)
        merge(ws, r, 6, r, 7)

    # IMPRESORA
    imp = data.get('impresora', {})
    wsection(ws, 'E35', 'IMPRESORA')
    merge(ws, 35, 5, 35, 7)
    ws['E36'] = 'MARCA';      ws['F36'] = str(imp.get('marca',''))
    ws['E37'] = 'REFERENCIA'; ws['F37'] = str(imp.get('referencia',''))
    ws['E38'] = 'SERIE';      ws['F38'] = str(imp.get('serie',''))
    ws['E39'] = 'IP';         ws['F39'] = str(imp.get('ip',''))
    for r in range(36, 40):
        for col in ['E','F']:
            ws[f'{col}{r}'].font = _font(size=11)
        merge(ws, r, 6, r, 7)

    # Observaciones
    ws['E41'] = 'Observaciones'
    ws['E41'].font = _font(bold=True, size=10)
    ws['F41'] = str(data.get('observaciones', ''))
    ws['F41'].font = _font(size=11)
    merge(ws, 41, 6, 41, 7)

    # ── Fila 42+: pie de página ───────────────────────────────────────────
    pie_row = max(col_izq_fin, 42)

    ws[f'B{pie_row}'] = 'BACKUP'
    ws[f'B{pie_row}'].font = _font(bold=True, size=13)
    ws[f'E{pie_row}'] = 'Responsable:'
    ws[f'E{pie_row}'].font = _font(bold=True, size=14)
    merge(ws, pie_row, 5, pie_row, 7)

    ws[f'B{pie_row+1}'] = 'ACTUALIZACIONES'
    ws[f'B{pie_row+1}'].font = _font(bold=True, size=13)
    ws[f'E{pie_row+1}'] = str(data.get('responsable', ''))
    ws[f'E{pie_row+1}'].font = _font(size=11)
    merge(ws, pie_row+1, 5, pie_row+1, 7)

    ws[f'B{pie_row+2}'] = 'ANTIVIRUS'
    ws[f'B{pie_row+2}'].font = _font(bold=True, size=13)

    # Ajustar alturas del pie
    for r in [pie_row, pie_row+1]:
        ws.row_dimensions[r].height = 18.75
    ws.row_dimensions[pie_row+2].height = 15


# ════════════════════════════════════════════════════════════════════════════
#  HOJA "Actualizaciones"
# ════════════════════════════════════════════════════════════════════════════
def build_actualizaciones(ws, novedades: list):
    """
    novedades: lista de dicts con keys:
      fecha, reportante, cedula, tipo_dano, descripcion, estado
    """
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20

    ws['A1'] = 'ACTUALIZACIONES / NOVEDADES'
    ws['A1'].font = _font(bold=True, size=13)

    headers = ['Fecha', 'Descripción / Tipo de Daño', 'Estado']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, size=11)
        c.alignment = _align(h='center')

    for ri, nov in enumerate(novedades, 3):
        fecha = nov.get('fecha_reporte', '')
        if hasattr(fecha, 'strftime'):
            fecha = fecha.strftime('%Y-%m-%d %H:%M')
        desc = f"[{nov.get('tipo_dano','')}] {nov.get('descripcion','')}"
        reportante = f"{nov.get('nombre_reportante','')} ({nov.get('cedula_reportante','')})"
        estado = nov.get('estado_ticket', '')

        row_data = [fecha, f"{desc}\nReportante: {reportante}", estado]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10)
            c.alignment = _align(wrap=True)
        ws.row_dimensions[ri].height = 36


# ════════════════════════════════════════════════════════════════════════════
#  FUNCIÓN PRINCIPAL: genera el .xlsx y lo escribe en disco
# ════════════════════════════════════════════════════════════════════════════
def generar_hoja_vida(data: dict, novedades: list, output_path: str):
    wb = openpyxl.Workbook()

    # Hoja inventario
    ws_inv = wb.active
    ws_inv.title = 'Inventario'
    build_inventario(ws_inv, data)

    # Hoja actualizaciones
    ws_act = wb.create_sheet('Actualizaciones')
    build_actualizaciones(ws_act, novedades)

    wb.save(output_path)
    return output_path


# ════════════════════════════════════════════════════════════════════════════
#  MODO CLI: llamado desde PHP → recibe JSON por stdin
#  php: $proc = proc_open('python3 gen_hoja_vida.py', ...)
#       fwrite($stdin, json_encode(['data'=>..., 'novedades'=>..., 'output'=>...]))
# ════════════════════════════════════════════════════════════════════════════
def _map_activo_a_data(activo: dict, hijos: list) -> dict:
    """Convierte el resultado de fun_read_activo_por_id al formato data del generador."""
    tipo = (activo.get('r_tipo') or '').lower()
    es_tablet = 'tablet' in tipo or 'ipad' in tipo

    # Clasificar hijos
    def get_hijo(tipo_str):
        for h in hijos:
            t = (h.get('r_tipo') or '').lower()
            if tipo_str in t:
                return h
        return {}

    def peri(h):
        return {
            'marca':     h.get('r_marca',''),
            'referencia':h.get('r_modelo','') or h.get('r_referencia',''),
            'serie':     h.get('r_serial',''),
        }

    lector_h  = get_hijo('lector')
    monitor_h = get_hijo('monitor') or get_hijo('pantalla')
    mouse_h   = get_hijo('mouse') or get_hijo('ratón') or get_hijo('raton')
    teclado_h = get_hijo('teclado') or get_hijo('keyboard')
    ups_h     = get_hijo('ups')
    imp_h     = get_hijo('impresora') or get_hijo('printer')
    tel_h     = get_hijo('teléfono') or get_hijo('telefono')
    tred_h    = get_hijo('tarjeta red') or get_hijo('inalambrica') or get_hijo('wifi')

    data = {
        'nombre_equipo':   activo.get('r_hostname') or activo.get('r_qr', ''),
        'usuario':         activo.get('r_responsable', ''),
        'dependencia':     activo.get('r_area', ''),
        'fecha':           datetime.today(),
        'responsable':     activo.get('r_responsable', ''),
        'tipo_raw':        activo.get('r_tipo', ''),
        'es_tablet':       es_tablet,
        'observaciones':   '',

        'tipo_equipo_label': (
            'Portatil _x_ Escritorio __ Todo en Uno __' if 'laptop' in tipo or 'portatil' in tipo
            else ('Tablet _x_' if es_tablet
            else 'Portatil __ Escritorio _x_ Todo en Uno __')
        ),

        'cpu': {
            'marca':         activo.get('r_marca', ''),
            'referencia':    activo.get('r_modelo', '') or activo.get('r_referencia', ''),
            'serial':        activo.get('r_serial', ''),
            'modelo':        activo.get('r_modelo', ''),
            'serial_envia':  '',
            'so':            '',
            'version':       '',
            'windows':       '',
            'service_pack':  '',
            'office':        '',
            'ip':            activo.get('r_ip', ''),
            'mac':           activo.get('r_mac', ''),
            'procesador':    '',
            'ram':           '',
            'disco_duro':    '',
            'cd':            '',
            'almacenamiento':'',
        },

        'lector':       peri(lector_h)  if lector_h  else {'marca':'NO APLICA','referencia':'','serie':''},
        'monitor':      peri(monitor_h) if monitor_h else {'marca':'NO APLICA','referencia':'','serie':''},
        'mouse':        peri(mouse_h)   if mouse_h   else {'marca':'NO APLICA','referencia':'','serie':''},
        'teclado':      peri(teclado_h) if teclado_h else {'marca':'NO APLICA','referencia':'','serie':''},
        'ups':          peri(ups_h)     if ups_h     else {'marca':'NO APLICA','referencia':'','serie':''},
        'impresora':    peri(imp_h)     if imp_h     else {'marca':'NO APLICA','referencia':'','serie':'', 'ip':''},
        'telefono':     peri(tel_h)     if tel_h     else {},
        'tarjeta_red':  {**peri(tred_h), 'mac': tred_h.get('r_mac','')} if tred_h else {},
        'base_lector':  {'marca':'','referencia':'','serie':'','serie_envia':''},
    }
    return data


# (ver bloque __main__ al final del archivo)


# ════════════════════════════════════════════════════════════════════════════
#  MODO GENERAL — una hoja por activo dentro del mismo libro
# ════════════════════════════════════════════════════════════════════════════
def generar_inventario_general(activos: list, output_path: str):
    wb = openpyxl.Workbook()
    first = True

    for act in activos:
        hijos     = act.get('_hijos', [])
        novedades = act.get('_novedades', [])
        data      = _map_activo_a_data(act, hijos)

        nombre_hoja = (act.get('r_hostname') or act.get('r_qr') or str(act.get('id_activo','')))
        nombre_hoja = nombre_hoja[:31]  # Excel limita a 31 chars

        if first:
            ws_inv = wb.active
            ws_inv.title = nombre_hoja
            first = False
        else:
            ws_inv = wb.create_sheet(nombre_hoja)

        build_inventario(ws_inv, data)

        if novedades:
            # Hoja de novedades junto a la de inventario
            ws_nov = wb.create_sheet(f"Nov_{nombre_hoja}"[:31])
            build_actualizaciones(ws_nov, novedades)

    wb.save(output_path)
    return output_path


# ── Sobreescribir el bloque __main__ para manejar ambos modos ───────────────
if __name__ == '__main__':
    import sys, json
    raw     = sys.stdin.read()
    payload = json.loads(raw)

    output  = payload['output']
    modo    = payload.get('modo', 'individual')

    if modo == 'general':
        generar_inventario_general(payload['activos'], output)
    else:
        activo    = payload['activo']
        hijos     = payload.get('hijos', [])
        novedades = payload.get('novedades', [])
        data      = _map_activo_a_data(activo, hijos)
        generar_hoja_vida(data, novedades, output)

    print(output)
